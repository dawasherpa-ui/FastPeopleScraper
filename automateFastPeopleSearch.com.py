import os
import time

from openpyxl import load_workbook
from playsound import playsound
from selenium import webdriver
# from selenium.webdriver.chrome.options import Options
import undetected_chromedriver as uc

import bs4


# VARIABLES
######################################################################################################################

chromedriver_path = "chromedriver.exe"  # Path to ChromeDriver executable
xlsx_path = "data.xlsx"  # Path to Excel file with names and addresses

CURR_SCRIPT_PATH = os.path.realpath(os.path.dirname(__file__))
# CURR_SCRIPT_PATH = os.path.dirname(sys.executable)
profile_path = CURR_SCRIPT_PATH + "\\profile"  # Path to Chrome profile (you can put the full path to existing profile or keep it to create new profile and use it later)

FIRST_NAME_COL = 'A'  # (input)
LAST_NAME_COL = 'B'  # (input)
ADDRESS_COL = 'K'  # (input)
MALING_COL='C'
PHONEs_COLs = ['L', 'M', 'N', 'O', 'P','Q']  # columns to output phone numbers  # (output)

######################################################################################################################

def open_chrome_with_profile():
    # Create a new Chrome session with the Chrome profile

    # options = Options()
    options = uc.ChromeOptions()
    options.add_argument("--user-data-dir=" + profile_path)

    # Create a new instance of the Chrome driver with the specified options
    # driver = webdriver.Chrome(executable_path=chromedriver_path, chrome_options=options)
    driver = uc.Chrome(driver_executable_path=chromedriver_path, options=options)
    return driver


def open_xlsx_file():
    # Open Excel file and return the workbook and worksheet

    wb = load_workbook(filename=xlsx_path)
    ws = wb.active
    return wb, ws


def write_phones_to_xlsx_file(wb, ws, phones,landNum,row):
    # Write up to three mobile phone numbers to Excel file (columns L, M, and N)
    for i in range(min(len(phones), 3)):
        ws.cell(row=row, column=12 + i, value=phones[i])  # Column L starts at index 12

    # Write landline numbers to Excel file (columns O, P, and Q) starting from the fourth column
    for i in range(min(len(landNum), 3)):
        ws.cell(row=row, column=15 + i, value=landNum[i])  # Column O starts at index 15

    wb.save(xlsx_path)

def extract_phones_from_page(page_source):
    # Extract phones from the page source and return them as a list of strings

    phones = []
    landNum=[]
    try:
        # find all phones
        soup = bs4.BeautifulSoup(page_source, "html.parser")
        phone_number_section = soup.find("div", id="phone_number_section")
        if phone_number_section:
            # find all a tags within the div
            col_elements = phone_number_section.find_all("dl", class_="col-sm-12 col-md-6")
            for col_element in col_elements:
                # Check if the word "wireless" is present in the dd element
                find_numBox = col_element.text.strip()
                # Flag to check if "wireless" is found in any dd element
                wireless_found = False
                if "Wireless" in find_numBox and "(Primary Phone)" in find_numBox:
                    wireless_found=True
                 # No need to continue checking other dd elements
                if wireless_found:
                    # If "wireless" is found in any dd element, find the associated a tag and extract its text
                    a_tag = col_element.find("a")
                    if a_tag:
                        phone_number = a_tag.text.strip()
                        phones.append(phone_number)

                else:
                    a_tag=col_element.find("a")
                    if a_tag:
                        phone_number = a_tag.text.strip()
                        landNum.append(phone_number)
        return phones,landNum

    except Exception as e:
        print(str(e))
        return phones


def main():
    driver = open_chrome_with_profile()  # Open Chrome with profile
    driver.get("https://www.fastpeoplesearch.com/")  # Navigate to FastPeopleSearch.com
    # time.sleep(80)
    # if access denied, wait for user to enable vpn (only for the first time)
    if "Access Denied" in driver.page_source:
        print("Access Denied")
        time.sleep(60)  # Wait for the user to enable vpn extension
        driver.get("https://www.fastpeoplesearch.com/")  # Navigate to FastPeopleSearch.com
        if "Access Denied" in driver.page_source:
            return 1
        
    wb, ws = open_xlsx_file()  # Open the Excel file
    # for each row in the Excel file search for the person and write the phones to the Excel file
    no_wireless=0
    for row in range(2, ws.max_row + 1):
        # try searching for this person
        try:
            print(row)
            first_name = ws[FIRST_NAME_COL + str(row)].value
            last_name = ws[LAST_NAME_COL + str(row)].value
            address = ws[ADDRESS_COL + str(row)].value
            address_maling = ws[MALING_COL + str(row)].value
            if (first_name is None and last_name is None) or address is None:
                continue

            # search for this person
            first_name = first_name.replace(" ", "-")
            last_name = last_name.replace(" ", "-")
            address = address.replace(" ", "-")
            driver.get("https://www.fastpeoplesearch.com/name/" + first_name + "-" + last_name + "_" + address)
            if "Are you human?" in driver.page_source:
                # Play a sound to alert the user
                playsound("notifier.mp3")
                # Pause the execution of the code for 80 seconds
                input("Please complete the captcha and press Enter to continue...")
            page_source = driver.page_source
            # Parse the HTML content using Beautiful Soup
            soup = bs4.BeautifulSoup(page_source, "html.parser")
            # Find all card elements
            card_elements = soup.find_all(class_="card")
            # Process each card element
            for card in card_elements:
                # Find phone numbers in the card
                phone_numbers = card.find_all("a", title=lambda x: x and "Search people with phone number" in x)
                address_check=card.text.strip()
                # If phone numbers are found
                if address_maling in address_check and phone_numbers and address_check != "Deceased":
                    print("Finding... "+first_name)
                    # Click on the card title element
                    card_title_element = card.find("h2", class_="card-title")
                    a_tag_element = card_title_element.find("a")
                    if a_tag_element:
                        href_attribute = a_tag_element.get("href")
                        # Click on the <a> tag element
                        driver.get("https://www.fastpeoplesearch.com"+href_attribute)
                        if "Are you human?" in driver.page_source:
                            # Play a sound to alert the user
                            playsound("notifier.mp3")
                            # Pause the execution of the code for 80 seconds
                            input("Please complete the captcha and press Enter to continue...")
                        profile_source=driver.page_source
                        # You can perform further actions after clicking on the <a> tag element
                        print("Extracting number....")
                        phones,landNum = extract_phones_from_page(profile_source)
                        if phones or landNum:
                            # Add your condition here
                            # For example, let's say you want to write phone data only if the person has more than 1 phone number
                            if len(phones) ==0:
                                no_wireless+1
                            if len(phones) > 0 or len(landNum) > 0:
                                # write phones to Excel file
                                print("Found " + str(len(phones)) + " wireless and "+str(len(landNum))+" landline number of " + first_name + " " + last_name)
                                write_phones_to_xlsx_file(wb, ws, phones,landNum, row)
                            else:
                                print(first_name," has only 0 phone number")
                        else:
                            print("No phones found for " + first_name + " " + last_name)
                        break
            print("Skiping..")
        except Exception as e:
            print(str(e))
            continue
    print(str(no_wireless)+ " wireless are empty")
    print("Finished")
    wb.close()
    driver.close()


if __name__ == "__main__":
    main()
